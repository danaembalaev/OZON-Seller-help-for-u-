from fastapi import FastAPI, UploadFile, File
from fastapi.responses import HTMLResponse, StreamingResponse
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

DEFAULT_STATUS = "Ожидает отгрузки"
SHEET_NAME = "Основной список"

app = FastAPI()

PAGE = """
<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>Ozon → Основной список</title>
  <style>
    body{font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial;max-width:760px;margin:40px auto;padding:0 16px}
    .card{border:1px solid #ddd;border-radius:16px;padding:18px}
    input,button{font-size:16px}
    button{padding:10px 14px;border-radius:12px;border:0;background:#111;color:#fff;cursor:pointer}
    .muted{color:#666;font-size:14px;line-height:1.4}
    .row{display:flex;gap:12px;align-items:center;flex-wrap:wrap}
    .pill{display:inline-block;padding:4px 10px;border-radius:999px;background:#f3f3f3;font-size:13px}
  </style>
</head>
<body>
  <h1>Загрузка отчёта → «Основной список» (Excel под печать)</h1>
  <div class="card">
    <div class="row">
      <span class="pill">Форматы: CSV / Excel</span>
      <span class="pill">Статус: Ожидает отгрузки</span>
    </div>
    <p class="muted">
      Можно загрузить <b>CSV</b> или <b>Excel</b> (xlsx/xls).
      Сервис сам попробует распознать нужные колонки даже если они названы чуть иначе.
    </p>
    <form action="/convert" method="post" enctype="multipart/form-data">
      <p><input type="file" name="file" required></p>
      <p><button type="submit">Скачать Excel</button></p>
    </form>
  </div>
</body>
</html>
"""

def read_csv_smart(content: bytes) -> pd.DataFrame:
    for enc in ("utf-8", "cp1251"):
        try:
            return pd.read_csv(BytesIO(content), sep=";", encoding=enc, engine="python")
        except Exception:
            continue
    return pd.read_csv(BytesIO(content), sep=";", engine="python")

def read_table_smart(filename: str, content: bytes) -> pd.DataFrame:
    name = (filename or "").lower()

    if name.endswith(".csv"):
        return read_csv_smart(content)

    if name.endswith(".xlsx") or name.endswith(".xls"):
        bio = BytesIO(content)
        return pd.read_excel(bio)

    try:
        return read_csv_smart(content)
    except Exception:
        raise ValueError("Неподдерживаемый формат. Загрузите CSV или Excel (.xlsx/.xls).")

def _norm(s: str) -> str:
    s = (s or "").strip().lower().replace("ё", "е")
    s = re.sub(r"[\s\-_]+", " ", s)
    s = re.sub(r"[^0-9a-zа-я ]+", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _pick_col(df: pd.DataFrame, variants: list[str]) -> str | None:
    norm_map = {_norm(c): c for c in df.columns}
    for v in variants:
        nv = _norm(v)
        if nv in norm_map:
            return norm_map[nv]
    for c in df.columns:
        nc = _norm(c)
        for v in variants:
            nv = _norm(v)
            if nv and nv in nc:
                return c
    return None

def ensure_required_columns(df: pd.DataFrame) -> pd.DataFrame:
    status_variants = ["Статус", "Статус заказа", "State", "Status"]
    article_variants = ["Артикул", "Артикул продавца", "Артикул товара", "Vendor code", "Vendor", "Offer id", "Offer_id"]
    qty_variants = ["Количество", "Кол-во", "Кол во", "Кол-во товара", "Qty", "Quantity"]
    ship_variants = ["Номер отправления", "Отправление", "Номер постинга", "Posting number", "Shipment", "Shipment number"]

    c_status = _pick_col(df, status_variants)
    c_art = _pick_col(df, article_variants)
    c_qty = _pick_col(df, qty_variants)
    c_ship = _pick_col(df, ship_variants)

    missing = []
    if not c_status: missing.append("Статус")
    if not c_art: missing.append("Артикул")
    if not c_qty: missing.append("Количество")
    if not c_ship: missing.append("Номер отправления")

    if missing:
        raise ValueError("В файле не найдены нужные колонки: " + ", ".join(missing))

    return df.rename(columns={
        c_status: "Статус",
        c_art: "Артикул",
        c_qty: "Количество",
        c_ship: "Номер отправления",
    }).copy()

def build_note(qtys: pd.Series) -> str:
    bigger = sorted({int(q) for q in qtys.dropna() if int(q) > 1})
    if not bigger:
        return ""
    return ", ".join([f"{q}в одну" for q in bigger])

def build_main_df(df_ready: pd.DataFrame) -> pd.DataFrame:
    grouped = (
        df_ready.groupby(["Артикул"], as_index=False)
        .agg(Итого=("Количество","sum"), Пометка=("Количество", build_note))
    )
    grouped["Количество"] = grouped.apply(
        lambda r: f"{int(r['Итого'])}({r['Пометка']})" if r["Пометка"] else f"{int(r['Итого'])}",
        axis=1
    )
    top_df = pd.DataFrame({
        "Артикул": grouped["Артикул"],
        "Количество": grouped["Количество"],
        "КОД": ["" for _ in range(len(grouped))]
    })

    blocks = []
    for ship, ship_df in df_ready.groupby("Номер отправления"):
        if ship_df["Артикул"].nunique() <= 1:
            continue
        items = (
            ship_df.groupby("Артикул")["Количество"]
            .sum()
            .reset_index()
            .sort_values("Артикул")
        )
        line = "; ".join([f"{r['Артикул']} — {int(r['Количество'])}" for _, r in items.iterrows()])
        blocks.append({"Артикул": line, "Количество": "", "КОД": ""})

    bottom_df = pd.DataFrame(blocks)
    spacer = pd.DataFrame([{"Артикул":"", "Количество":"", "КОД":""} for _ in range(2)])
    return pd.concat([top_df, spacer, bottom_df], ignore_index=True)

def format_as_osnovnoi_spisok(xlsx_bytes: bytes) -> bytes:
    bio = BytesIO(xlsx_bytes)
    wb = load_workbook(bio)
    ws = wb.active
    ws.title = SHEET_NAME

    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left = Alignment(horizontal="left", vertical="center", wrap_text=False)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            if cell.column == 2:
                cell.alignment = center
            else:
                cell.alignment = left

    widths = {1: 30, 2: 18, 3: 15}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        if (isinstance(a, str) and a.strip() and (not b) and (not c) and ("—" in a or ";" in a)):
            ws.cell(row=r, column=1).font = Font(bold=True)

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75

    out = BytesIO()
    wb.save(out)
    return out.getvalue()

@app.get("/", response_class=HTMLResponse)
def index():
    return PAGE

@app.post("/convert")
async def convert(file: UploadFile = File(...)):
    content = await file.read()

    try:
        df = read_table_smart(file.filename, content)
        df = ensure_required_columns(df)
    except ValueError as e:
        return HTMLResponse(f"<h3>{e}</h3>", status_code=400)

    df_ready = df[df["Статус"] == DEFAULT_STATUS].copy()
    df_ready["Количество"] = pd.to_numeric(df_ready["Количество"], errors="coerce").fillna(0).astype(int)

    final_df = build_main_df(df_ready)

    xlsx_io = BytesIO()
    final_df.to_excel(xlsx_io, index=False)
    formatted = format_as_osnovnoi_spisok(xlsx_io.getvalue())

    filename = "Osnovnoi_spisok.xlsx"
    return StreamingResponse(
        BytesIO(formatted),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
